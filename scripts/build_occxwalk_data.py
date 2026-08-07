"""Build occxwalk's embedded Stata lookup datasets from the source workbooks."""

from __future__ import annotations

import argparse
from collections import defaultdict
from hashlib import sha256
import json
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


TEXT_CODE_SYSTEMS = {"ONET_SOC2019_full", "SOC2010"}
SYSTEMS = [
    "CFPS",
    "CGSS06",
    "CSS",
    "GB2015_full",
    "GB2015_reduce",
    "GB2022",
    "GB9909",
    "ISCO08",
    "ISCO68",
    "ISCO88",
    "ONET_SOC2019_full",
    "SOC2010",
]
FILENAME_PATTERN = re.compile(r"职业体系匹配_主体系_(.+)_GPT56\.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build occxwalk_catalog.dta and occxwalk_links.dta."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        required=True,
        help="Directory containing the 12 职业体系匹配_主体系_*_GPT56.xlsx files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Destination directory (default: repository root).",
    )
    parser.add_argument(
        "--data-version",
        default="GPT56",
        help="Version string written to occxwalk_manifest.json.",
    )
    return parser.parse_args()


def workbook_system(path: Path) -> str:
    match = FILENAME_PATTERN.fullmatch(path.name)
    if not match:
        raise ValueError(f"Unexpected workbook filename: {path.name}")
    return match.group(1)


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalized_key(system: str, code: object) -> str:
    value = clean_text(code)
    if not value:
        raise ValueError(f"Blank code in {system}")
    if system in TEXT_CODE_SYSTEMS:
        return value.upper()
    if not value.isdigit():
        raise ValueError(f"Non-numeric code {value!r} in numeric system {system}")
    return str(int(value))


def file_hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_workbooks(input_dir: Path) -> tuple[dict[str, list[dict[str, object]]], list[dict[str, object]]]:
    workbook_rows: dict[str, list[dict[str, object]]] = {}
    source_manifest: list[dict[str, object]] = []
    paths = sorted(input_dir.glob("职业体系匹配_主体系_*_GPT56.xlsx"))

    for path in paths:
        system = workbook_system(path)
        if system not in SYSTEMS:
            raise ValueError(f"Unrecognized system in filename: {system}")
        if system in workbook_rows:
            raise ValueError(f"More than one workbook found for {system}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        if "匹配结果" not in workbook.sheetnames:
            raise ValueError(f"Sheet 匹配结果 not found in {path.name}")
        rows = workbook["匹配结果"].iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows)]
        required = {
            f"{system}_代码",
            f"{system}_名称",
            f"{system}_描述",
            f"{system}_源行号",
        }
        for target in SYSTEMS:
            if target != system:
                required.update(
                    {f"{target}_代码", f"{target}_名称", f"{target}_置信度"}
                )
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(f"Missing columns in {path.name}: {missing}")

        records = [dict(zip(headers, row, strict=True)) for row in rows]
        workbook_rows[system] = records
        source_manifest.append(
            {
                "file": path.name,
                "system": system,
                "rows": len(records),
                "sha256": file_hash(path),
            }
        )

    missing_systems = sorted(set(SYSTEMS) - set(workbook_rows))
    if missing_systems:
        raise ValueError(f"Missing source workbooks for: {missing_systems}")
    return workbook_rows, source_manifest


def build_tables(
    workbook_rows: dict[str, list[dict[str, object]]],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, list[str]], dict[str, int]]:
    catalog_rows: list[dict[str, object]] = []
    link_rows: list[dict[str, object]] = []
    ambiguous_keys: dict[str, list[str]] = defaultdict(list)
    exact_duplicate_counts: dict[str, int] = defaultdict(int)

    for system in SYSTEMS:
        grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
        for record in workbook_rows[system]:
            grouped[normalized_key(system, record[f"{system}_代码"])].append(record)

        for key, variants in grouped.items():
            variants.sort(key=lambda record: int(record[f"{system}_源行号"]))
            chosen = variants[0]
            comparison_fields = [field for field in chosen if not field.endswith("_源行号")]
            signatures = {
                tuple(clean_text(record[field]) for field in comparison_fields)
                for record in variants
            }
            ambiguous = int(len(signatures) > 1)
            if ambiguous:
                ambiguous_keys[system].append(clean_text(chosen[f"{system}_代码"]))
            elif len(variants) > 1:
                exact_duplicate_counts[system] += len(variants) - 1

            catalog_rows.append(
                {
                    "system": system,
                    "key": key,
                    "code": clean_text(chosen[f"{system}_代码"]),
                    "name": clean_text(chosen[f"{system}_名称"]),
                    "description": clean_text(chosen[f"{system}_描述"]),
                    "ambiguous": ambiguous,
                    "source_row": int(chosen[f"{system}_源行号"]),
                }
            )
            for target in SYSTEMS:
                if target == system:
                    continue
                link_rows.append(
                    {
                        "from_system": system,
                        "from_key": key,
                        "to_system": target,
                        "to_code": clean_text(chosen[f"{target}_代码"]),
                        "to_name": clean_text(chosen[f"{target}_名称"]),
                        "confidence": float(chosen[f"{target}_置信度"]),
                        "ambiguous": ambiguous,
                        "source_row": int(chosen[f"{system}_源行号"]),
                    }
                )

    catalog = pd.DataFrame(catalog_rows).sort_values(
        ["system", "key"], kind="stable"
    ).reset_index(drop=True)
    links = pd.DataFrame(link_rows).sort_values(
        ["from_system", "from_key", "to_system"], kind="stable"
    ).reset_index(drop=True)
    catalog = catalog.astype({"ambiguous": "int8", "source_row": "int32"})
    links = links.astype(
        {"ambiguous": "int8", "source_row": "int32", "confidence": "float64"}
    )

    if catalog.duplicated(["system", "key"]).any():
        raise ValueError("Catalog key is not unique")
    if links.duplicated(["from_system", "from_key", "to_system"]).any():
        raise ValueError("Link key is not unique")
    if catalog[["system", "key", "code", "name", "description"]].isna().any().any():
        raise ValueError("Catalog contains missing required values")
    if links[["from_system", "from_key", "to_system", "to_code", "to_name", "confidence"]].isna().any().any():
        raise ValueError("Links contain missing required values")
    expected_links = len(catalog) * (len(SYSTEMS) - 1)
    if len(links) != expected_links:
        raise ValueError(f"Expected {expected_links} links, got {len(links)}")
    if not links["confidence"].between(0, 1).all():
        raise ValueError("Confidence outside [0, 1]")
    return catalog, links, dict(ambiguous_keys), dict(exact_duplicate_counts)


def write_outputs(
    output_dir: Path,
    data_version: str,
    catalog: pd.DataFrame,
    links: pd.DataFrame,
    ambiguous_codes: dict[str, list[str]],
    exact_duplicates: dict[str, int],
    source_manifest: list[dict[str, object]],
) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    catalog.to_stata(
        output_dir / "occxwalk_catalog.dta",
        write_index=False,
        version=118,
        convert_strl=["description"],
        variable_labels={
            "system": "Canonical occupation coding system",
            "key": "Normalized source-code merge key",
            "code": "Original occupation code",
            "name": "Occupation name",
            "description": "Occupation description",
            "ambiguous": "1 if duplicated code has conflicting source rows",
            "source_row": "Selected Excel source row",
        },
    )
    links.to_stata(
        output_dir / "occxwalk_links.dta",
        write_index=False,
        version=118,
        variable_labels={
            "from_system": "Source occupation coding system",
            "from_key": "Normalized source-code merge key",
            "to_system": "Target occupation coding system",
            "to_code": "Matched target occupation code",
            "to_name": "Matched target occupation name",
            "confidence": "GPT56 match confidence",
            "ambiguous": "1 if source code has conflicting source rows",
            "source_row": "Selected Excel source row",
        },
    )
    manifest: dict[str, object] = {
        "package": "occxwalk",
        "data_version": data_version,
        "systems": SYSTEMS,
        "text_code_systems": sorted(TEXT_CODE_SYSTEMS),
        "catalog_rows": len(catalog),
        "link_rows": len(links),
        "ambiguous_codes": {
            system: sorted(codes) for system, codes in ambiguous_codes.items()
        },
        "exact_duplicate_rows_removed": exact_duplicates,
        "source_workbooks": source_manifest,
    }
    (output_dir / "occxwalk_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def main() -> None:
    args = parse_args()
    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    workbook_rows, source_manifest = read_workbooks(input_dir)
    catalog, links, ambiguous, duplicates = build_tables(workbook_rows)
    manifest = write_outputs(
        output_dir,
        args.data_version,
        catalog,
        links,
        ambiguous,
        duplicates,
        source_manifest,
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
